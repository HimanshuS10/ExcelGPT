import openpyxl

def modify_excel(file_path, instruction: str):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Example: Handle "add A1 and B1 into A3"
    if "add" in instruction.lower():
        parts = instruction.split()
        try:
            # crude parsing — better if Gemini gives JSON later
            cell1, cell2, target = "A1", "B1", "A3"
            ws[target].value = ws[cell1].value + ws[cell2].value
        except Exception as e:
            return f"Error: {str(e)}"

    wb.save(file_path)
    return file_path
